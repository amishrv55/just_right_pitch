# proposals/views.py
from django.shortcuts import render, redirect
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.contrib.auth.models import User

from django.db.models.functions import TruncMonth, ExtractWeekDay
import json

from .forms import SignUpForm, ProfileForm
from .models import Profile, Proposal
# proposals/views.py (append these imports at top if missing)
from django.views.decorators.http import require_http_methods
from .forms import ProposalForm
from .utils import fetch_job_description
from .models import Proposal, Profile, Contract, Payment, Client
from .prompts import PLATFORM_PROMPTS

from openai import OpenAI
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from django.shortcuts import get_object_or_404
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.units import inch
from django.db.models import Count, Q, Avg, F, Sum
from django.utils.timezone import now
from collections import Counter
import calendar
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from django.template.loader import render_to_string
import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import io
import csv
import openpyxl
import io, zipfile
from .models import FreelancerDirectoryProfile
from .forms import FreelancerDirectoryProfileForm
from reportlab.lib import colors
from .utils_reminders import get_user_reminders

from django.contrib.auth.decorators import user_passes_test
from django.utils.timezone import now

from proposals.models import Profile, CreditTransaction
from proposals.forms import AdminAdjustCreditsForm, CreditRequestForm
from proposals.services.credits import adjust_credits

from django.db import transaction

from .models import Conversation
from io import BytesIO

from django.contrib.admin.views.decorators import staff_member_required

COST_PER_GENERATION = 1  # credits per AI proposal


client = OpenAI(api_key=settings.OPENAI_API_KEY)

#def _build_user_prompt(job_desc: str, profile: Profile, tone: str) -> str:
 #   return f"""
#Job Description:
#{job_desc}

#Freelancer Profile (Writer of Proposal):
#Name: {profile.full_name or profile.user.username}
#Skills: {profile.skills or "-"}
#Portfolio: {profile.portfolio or "-"}

#Tone: {tone}
#""".strip()

def _build_user_prompt(job_description, profile, tone):
    """
    Builds the user prompt for the AI model including name from profile.
    """
    name = profile.full_name or profile.user.get_full_name() or profile.user.username

    return f"""
Job Description:
{job_description}

Write a {tone} proposal for the above job.  
End the proposal with:

Best regards,
{name}
"""

@login_required
@require_http_methods(["GET", "POST"])

def create_proposal(request):
    profile, _ = Profile.objects.get_or_create(user=request.user)
    initial = {"tone": profile.preferred_tone}

    if request.method == "POST":
        action = request.POST.get("action")

        # Manual proposal
        if action == "manual":
            platform = request.POST.get("platform", "Other")
            tone = request.POST.get("tone", "Professional")
            manual_text = request.POST.get("manual_proposal_text", "").strip()
            job_title = request.POST.get("job_title", "").strip()
            job_description_text = request.POST.get("job_description_text", "").strip()

            if not manual_text:
                messages.warning(request, "Please write your proposal before saving.")
                return render(request, "proposals/create_proposal.html", {"form": ProposalForm(initial=initial)})

            new_proposal = Proposal.objects.create(
                user=request.user,
                platform=platform,
                tone=tone,
                job_title=job_title,
                job_description=job_description_text,
                proposal_text=manual_text,
            )
            messages.success(request, "Proposal saved successfully.")
            return redirect("proposal_detail", proposal_id=new_proposal.id)

        # AI proposal (generate or fetch flow)
        form = ProposalForm(request.POST)
        if form.is_valid():
            platform = form.cleaned_data["platform"]
            tone = form.cleaned_data["tone"] or profile.preferred_tone
            job_title = form.cleaned_data["job_title"]
            job_description = form.cleaned_data["job_description"]
            job_url = form.cleaned_data.get("job_url")

            # --- FETCH MODE ---
            if action == "fetch":
                if job_url:
                    desc, err = fetch_job_description(job_url)
                    if err:
                        messages.error(request, f"Failed to import: {err}")
                    elif desc:
                        messages.success(request, "Job description imported successfully.")
                        form = ProposalForm(initial={"platform": platform, "tone": tone, "job_description": desc})
                else:
                    messages.warning(request, "Please provide a valid job URL to fetch.")
                return render(request, "proposals/create_proposal.html", {"form": form})

            # --- GENERATE MODE ---
            if action == "generate":
                if not job_description.strip():
                    messages.warning(request, "Please paste or import a job description before generating.")
                    return render(request, "proposals/create_proposal.html", {"form": form})

                # ✅ Step 1: Check credits
                profile.refresh_from_db()
                if profile.ai_credits < COST_PER_GENERATION:
                    messages.error(request, "Not enough AI credits. Please request a top-up before generating.")
                    return render(request, "proposals/create_proposal.html", {"form": form})

                system_prompt = PLATFORM_PROMPTS.get(platform, PLATFORM_PROMPTS["Generic"])
                user_prompt = _build_user_prompt(job_description, profile, tone)

                try:
                    resp = client.chat.completions.create(
                        model="gpt-3.5-turbo",
                        temperature=0.7,
                        messages=[
                            {"role": "system", "content": system_prompt},
                            {"role": "user", "content": user_prompt},
                        ],
                    )
                    proposal_text = resp.choices[0].message.content.strip() if resp.choices else ""
                    proposal_text = proposal_text.replace("[Your Name]", profile.full_name or profile.user.username)
                except Exception as e:
                    messages.error(request, f"AI generation error: {e}")
                    return render(request, "proposals/create_proposal.html", {"form": form})

                if not proposal_text:
                    messages.error(request, "Generation returned empty text. Please try again.")
                    return render(request, "proposals/create_proposal.html", {"form": form})

                # ✅ Step 2: Save proposal
                new_proposal = Proposal.objects.create(
                    user=request.user,
                    job_title=job_title,
                    job_description=job_description,
                    platform=platform,
                    tone=tone,
                    proposal_text=proposal_text,
                )

                # ✅ Step 3: Deduct credits (safe, atomic)
                try:
                    adjust_credits(
                        request.user,
                        -COST_PER_GENERATION,
                        reason="generation",
                        note=f"Proposal #{new_proposal.id}"
                    )
                except ValueError:
                    # rollback in rare race condition
                    new_proposal.delete()
                    messages.error(request, "Credits changed unexpectedly. Please top-up and try again.")
                    return render(request, "proposals/create_proposal.html", {"form": form})

                messages.success(request, "Proposal generated and saved. 1 credit deducted.")
                return redirect("proposal_detail", proposal_id=new_proposal.id)

        messages.error(request, "Please correct the errors below.")
        return render(request, "proposals/create_proposal.html", {"form": form})

    # GET request
    form = ProposalForm(initial=initial)
    return render(request, "proposals/create_proposal.html", {"form": form})


@login_required
def proposal_detail(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)

    if request.method == "POST":
        if request.POST.get("action") == "save":
            proposal.proposal_text = request.POST.get("proposal_text", proposal.proposal_text)
            proposal.save()
            messages.success(request, "Proposal updated successfully.")
            return redirect("proposal_detail", proposal_id=proposal.id)

    return render(request, "proposals/proposal_detail.html", {"proposal": proposal})


# --- Signup ---
def signup(request):
    if request.user.is_authenticated:
        return redirect("dashboard")

    if request.method == "POST":
        form = SignUpForm(request.POST)
        if form.is_valid():
            user = form.save()
            # Create a blank profile
            Profile.objects.get_or_create(user=user)
            login(request, user)  # Auto login after signup
            return redirect("dashboard")
    else:
        form = SignUpForm()

    return render(request, "proposals/signup.html", {"form": form})


# --- Dashboard ---

from proposals.scheduler import scheduler  # ✅ import your scheduler instance

@login_required
def dashboard(request):
    search_query = request.GET.get("search", "").strip()
    platform_filter = request.GET.get("platform", "")
    status_filter = request.GET.get("status", "")

    # Fetch proposals with related contracts for efficiency
    proposals = Proposal.objects.filter(user=request.user).select_related("contract").order_by("-created_at")
    proposal_alerts, payment_alerts = get_user_reminders(request.user)

    # === Filtering logic ===
    if search_query:
        proposals = proposals.filter(
            Q(proposal_text__icontains=search_query) |
            Q(job_description__icontains=search_query) |
            Q(note__icontains=search_query)
        )
    if platform_filter and platform_filter != "ALL":
        proposals = proposals.filter(platform=platform_filter)
    if status_filter and status_filter != "ALL":
        proposals = proposals.filter(status=status_filter)

    # Collect distinct filters
    platforms = Proposal.objects.filter(user=request.user).values_list("platform", flat=True).distinct()
    statuses = Proposal.objects.filter(user=request.user).values_list("status", flat=True).distinct()

    # === Quick metrics ===
    won_count = proposals.filter(status="Won").count()
    pending_count = proposals.filter(status="Sent").count()

    # === Scheduler job info ===
    scheduled_jobs = []
    if scheduler:
        try:
            for job in scheduler.get_jobs():
                scheduled_jobs.append({"id": job.id, "next_run": job.next_run_time})
        except Exception as e:
            # If scheduler hasn't started, just skip silently
            print(f"⚠️ Could not fetch jobs: {e}")

    # === Choices for dropdowns ===
    status_choices = Proposal._meta.get_field("status").choices
    confidence_choices = Proposal._meta.get_field("confidence").choices

    context = {
        "proposals": proposals,
        "platforms": platforms,
        "statuses": statuses,
        "status_choices": status_choices,
        "confidence_choices": confidence_choices,
        "search_query": search_query,
        "platform_filter": platform_filter,
        "status_filter": status_filter,
        "proposal_alerts": proposal_alerts,
        "payment_alerts": payment_alerts,
        "won_count": won_count,
        "pending_count": pending_count,
        "scheduled_jobs": scheduled_jobs,  # ✅ added to context
    }
    return render(request, "proposals/dashboard.html", context)

# --- Profile ---
@login_required
def profile(request):
    profile, created = Profile.objects.get_or_create(user=request.user)
    if request.method == "POST":
        form = ProfileForm(request.POST, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Profile updated successfully.")
            return redirect("profile")
    else:
        form = ProfileForm(instance=profile)

    return render(request, "proposals/profile.html", {"form": form})


# --- Logout ---
def user_logout(request):
    logout(request)
    return redirect(settings.LOGOUT_REDIRECT_URL)


@login_required
def download_proposal_pdf(request, proposal_id):
    proposal = Proposal.objects.get(id=proposal_id, user=request.user)

    # PDF HTTP response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="proposal_{proposal_id}.pdf"'

    # PDF document
    doc = SimpleDocTemplate(response, pagesize=letter, rightMargin=50, leftMargin=50, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        name="TitleStyle",
        parent=styles["Heading1"],
        alignment=TA_CENTER,
        fontSize=20,
        spaceAfter=20
    )
    section_title_style = ParagraphStyle(
        name="SectionTitle",
        parent=styles["Heading2"],
        fontSize=14,
        textColor="#333333",
        spaceBefore=12,
        spaceAfter=6
    )
    content_style = ParagraphStyle(
        name="Content",
        parent=styles["Normal"],
        fontSize=11,
        leading=15,
        alignment=TA_LEFT
    )

    # PDF elements
    elements = []

    # Title
    elements.append(Paragraph("Just Right Pitch - Proposal", title_style))
    elements.append(Spacer(1, 12))

    # Platform and Tone
    elements.append(Paragraph(f"<b>Platform:</b> {proposal.platform}", section_title_style))
    elements.append(Paragraph(f"<b>Tone:</b> {proposal.tone}", content_style))
    elements.append(Spacer(1, 12))

    # Job Description
    if proposal.job_description:
        elements.append(Paragraph("Job Description", section_title_style))
        elements.append(Paragraph(proposal.job_description.replace("\n", "<br/>"), content_style))
        elements.append(Spacer(1, 12))

    # Proposal Text
    elements.append(Paragraph("Generated Proposal", section_title_style))
    elements.append(Paragraph(proposal.proposal_text.replace("\n", "<br/>"), content_style))

    # Build the PDF
    doc.build(elements)

    return response

@login_required
def delete_proposal(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    if request.method == "POST":
        proposal.delete()
        messages.success(request, "Proposal deleted successfully.")
        return redirect("dashboard")

    # If GET, show confirmation page (optional)
    return render(request, "proposals/confirm_delete.html", {"proposal": proposal})

@login_required
def update_proposal_tracking(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    if request.method == "POST":
        proposal.status = request.POST.get("status", proposal.status)
        proposal.confidence = request.POST.get("confidence", proposal.confidence)
        proposal.note = request.POST.get("note", proposal.note)
        proposal.save()
        messages.success(request, "Proposal tracking updated.")
    return redirect("dashboard")


@login_required
def analytics(request):
    proposals = Proposal.objects.filter(user=request.user)
    payments = Payment.objects.filter(proposal__user=request.user)

    # ✅ Status-based counts
    total_sent = proposals.filter(status__in=["Sent", "Won", "Lost"]).count()
    total_won = proposals.filter(status="Won").count()
    win_rate = (total_won / total_sent * 100) if total_sent > 0 else 0

    # ✅ Monthly proposal activity
    monthly_qs = (
        proposals.annotate(month=TruncMonth("created_at"))
        .values("month")
        .annotate(count=Count("id"))
        .order_by("month")
    )
    monthly_labels = [m["month"].strftime("%b %Y") for m in monthly_qs]
    monthly_counts = [m["count"] for m in monthly_qs]
    monthly_activity = {"labels": monthly_labels, "data": monthly_counts}

    # ✅ Wins by Confidence
    CONFIDENCE_LABELS = {
        "green": "High Confidence",
        "orange": "Medium Confidence",
        "red": "Low Confidence",
    }
    confidence_qs = (
        proposals.filter(status="Won")
        .values("confidence")
        .annotate(count=Count("id"))
    )
    confidence_labels = [
        CONFIDENCE_LABELS.get(c["confidence"].lower(), c["confidence"].capitalize())
        for c in confidence_qs
    ]
    confidence_counts = [c["count"] for c in confidence_qs]
    confidence_wins = {"labels": confidence_labels, "data": confidence_counts}

    # ✅ Platform performance
    platform_qs = (
        proposals.filter(status="Won")
        .values("platform")
        .annotate(count=Count("id"))
    )
    platform_labels = [p["platform"] for p in platform_qs]
    platform_counts = [p["count"] for p in platform_qs]
    platform_performance = {"labels": platform_labels, "data": platform_counts}

    # ✅ Best platform
    best_platform = platform_qs.order_by("-count").first()
    best_platform_name = best_platform["platform"] if best_platform else "N/A"

    # ✅ Average contract/payment value
    avg_value = payments.aggregate(avg=Avg("amount"))["avg"] or 0

    # ✅ Day-of-week activity
    dow_qs = (
        proposals.annotate(dow=ExtractWeekDay("created_at"))
        .values("dow")
        .annotate(count=Count("id"))
        .order_by("dow")
    )
    dow_labels = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]
    dow_data = [0] * 7
    for d in dow_qs:
        dow_data[d["dow"] - 1] = d["count"]
    day_of_week_activity = {"labels": dow_labels, "data": dow_data}

    # ✅ Top clients
    client_qs = (
        proposals.filter(client__isnull=False)
        .values("client__name")
        .annotate(count=Count("id"))
        .order_by("-count")[:5]
    )
    top_clients = {
        "labels": [c["client__name"] for c in client_qs],
        "data": [c["count"] for c in client_qs],
    }

    # ✅ Outstanding payments
    outstanding_qs = (
        payments.filter(status="Pending")
        .values("proposal__client__name")
        .annotate(total=Sum("amount"))
    )
    outstanding_payments = {
        "labels": [
            o["proposal__client__name"] or "Unknown Client" for o in outstanding_qs
        ],
        "data": [float(o["total"]) for o in outstanding_qs],
    }

    return render(request, "proposals/analytics.html", {
        "total_sent": total_sent,
        "total_won": total_won,
        "win_rate": round(win_rate, 1),
        "best_platform": best_platform_name,
        "avg_value": round(avg_value, 2),

        # JSON for charts
        "monthly_activity": json.dumps(monthly_activity),
        "confidence_wins": json.dumps(confidence_wins),
        "platform_performance": json.dumps(platform_performance),
        "day_of_week_activity": json.dumps(day_of_week_activity),
        "top_clients": json.dumps(top_clients),
        "outstanding_payments": json.dumps(outstanding_payments),
    })

@login_required
def generate_contract(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    contract_text = f"""
    Contract Agreement
    -------------------
    This agreement is made between {request.user.profile.full_name} and the client.
    Project: {proposal.job_description[:50]}...
    Payment Terms: As agreed
    Deliverables: As agreed
    """
    contract = Contract.objects.create(proposal=proposal, contract_text=contract_text)
    messages.success(request, "Contract generated successfully.")
    return redirect("view_contract", contract.id)

@login_required
def view_contract(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    contract = Contract.objects.filter(proposal=proposal).first()
    if not contract:
        messages.warning(request, "No contract found for this proposal.")
        return redirect('dashboard')
    return render(request, 'proposals/view_contract.html', {'contract': contract})

@login_required
def add_payment(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    
    if request.method == "POST":
        amount = request.POST.get("amount")
        status = request.POST.get("status")
        due_date = request.POST.get("due_date")
        received_date = request.POST.get("received_date") or None

        # Create the payment record
        Payment.objects.create(
            proposal=proposal,
            amount=amount,
            status=status,
            due_date=due_date if due_date else None,
            received_date=received_date
        )

        # If payment is received, update proposal status automatically
        if status == "received":
            proposal.status = "Won"  # Mark project as won if payment received
            proposal.save(update_fields=["status"])

        messages.success(request, "Payment details saved successfully.")
        return redirect("dashboard")

    return render(request, "proposals/add_payment.html", {"proposal": proposal})


@login_required
@csrf_exempt
def update_proposal_status(request, proposal_id):
    if request.method == "POST":
        proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
        status = request.POST.get("status")
        if status in dict(Proposal.STATUS_CHOICES):
            proposal.status = status
            proposal.save()
            return JsonResponse({"success": True, "status": status})
        return JsonResponse({"success": False, "error": "Invalid status"}, status=400)
    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

@login_required
@csrf_exempt
def update_proposal_confidence(request, proposal_id):
    if request.method == "POST":
        proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
        confidence = request.POST.get("confidence")
        if confidence in dict(Proposal.CONFIDENCE_CHOICES):
            proposal.confidence = confidence
            proposal.save()
            return JsonResponse({"success": True, "confidence": confidence})
        return JsonResponse({"success": False, "error": "Invalid confidence"}, status=400)
    return JsonResponse({"success": False, "error": "Invalid request"}, status=400)

@login_required
def track_payments(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    payments = Payment.objects.filter(proposal=proposal).order_by("due_date", "received_date")

    return render(request, "proposals/payments_track.html", {
        "proposal": proposal,
        "payments": payments
    })

@login_required
def contract_form(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    return render(request, "proposals/contract_form.html", {"proposal": proposal})


@login_required
def generate_contract_pdf_simple(request):
    if request.method == "POST":
        proposal_id = request.POST.get("proposal_id")
        first_party = request.POST.get("first_party")
        second_party = request.POST.get("second_party")
        effective_date = request.POST.get("effective_date")
        terms = request.POST.get("terms")

        proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)

        # Create PDF HTTP response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="contract_{proposal.id}.pdf"'

        # Create PDF document
        doc = SimpleDocTemplate(
            response,
            pagesize=letter,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50
        )

        styles = getSampleStyleSheet()

        # Custom styles
        title_style = ParagraphStyle(
            name="TitleStyle",
            parent=styles["Heading1"],
            alignment=TA_CENTER,
            fontSize=20,
            spaceAfter=20
        )
        section_title_style = ParagraphStyle(
            name="SectionTitle",
            parent=styles["Heading2"],
            fontSize=14,
            textColor="#333333",
            spaceBefore=12,
            spaceAfter=6
        )
        content_style = ParagraphStyle(
            name="Content",
            parent=styles["Normal"],
            fontSize=11,
            leading=15,
            alignment=TA_LEFT
        )

        elements = []

        # Title
        elements.append(Paragraph("Service Agreement", title_style))
        elements.append(Spacer(1, 12))

        # Parties
        elements.append(Paragraph("Parties", section_title_style))
        elements.append(Paragraph(f"First Party: {first_party}", content_style))
        elements.append(Paragraph(f"Second Party: {second_party}", content_style))
        elements.append(Spacer(1, 12))

        # Effective Date
        elements.append(Paragraph("Effective Date", section_title_style))
        elements.append(Paragraph(effective_date, content_style))
        elements.append(Spacer(1, 12))

        # Project Details
        elements.append(Paragraph("Project Details", section_title_style))
        elements.append(Paragraph(proposal.job_description.replace("\n", "<br/>"), content_style))
        elements.append(Spacer(1, 12))

        # Terms & Conditions
        elements.append(Paragraph("Terms & Conditions", section_title_style))
        elements.append(Paragraph(terms.replace("\n", "<br/>"), content_style))
        elements.append(Spacer(1, 12))

        # Signatures
        elements.append(Paragraph("Signatures", section_title_style))
        elements.append(Paragraph("First Party: _______________________", content_style))
        elements.append(Paragraph("Second Party: ______________________", content_style))
        elements.append(Spacer(1, 20))

        # Build PDF
        doc.build(elements)
        return response


@login_required
def add_client_to_proposal(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)

    if request.method == "POST":
        if "existing_client" in request.POST and request.POST.get("existing_client"):
            # Attach existing client
            client_id = request.POST.get("existing_client")
            client = get_object_or_404(Client, id=client_id, user=request.user)
            proposal.client = client
            proposal.save()
            messages.success(request, f"Client {client.name} assigned successfully to this proposal.")
            return redirect("dashboard")

        else:
            # Create new client
            name = request.POST.get("name")
            email = request.POST.get("email")
            notes = request.POST.get("notes")

            if name:
                client, created = Client.objects.get_or_create(
                    user=request.user,
                    name=name,
                    defaults={"email": email, "notes": notes}
                )
                proposal.client = client
                proposal.save()
                messages.success(request, f"Client {client.name} added and assigned successfully.")
                return redirect("dashboard")
            else:
                messages.error(request, "Client name is required to add a new client.")

    # Pass existing clients for dropdown
    existing_clients = Client.objects.filter(user=request.user)
    return render(request, "proposals/add_client.html", {"proposal": proposal, "existing_clients": existing_clients})

@login_required
def my_clients(request):
    clients = Client.objects.filter(user=request.user)

    # annotate number of projects linked to each client
    client_data = []
    for client in clients:
        project_count = Proposal.objects.filter(user=request.user, client=client).count()
        # revenue placeholder (we’ll add once payments have client link)
        revenue = 0  
        client_data.append({
            "id": client.id,
            "name": client.name,
            "email": client.email,
            "notes": client.notes,
            "projects": project_count,
            "revenue": revenue,
        })

    return render(request, "proposals/my_clients.html", {"clients": client_data})


@login_required
def edit_client(request, client_id):
    client = get_object_or_404(Client, id=client_id, user=request.user)

    if request.method == "POST":
        client.name = request.POST.get("name")
        client.email = request.POST.get("email")
        client.notes = request.POST.get("notes")
        client.save()
        messages.success(request, "Client updated successfully!")
        return redirect("my_clients")

    return render(request, "proposals/edit_client.html", {"client": client})


@login_required
def delete_client(request, client_id):
    client = get_object_or_404(Client, id=client_id, user=request.user)
    client.delete()
    messages.success(request, "Client deleted successfully!")
    return redirect("my_clients")


# -------------------------
# Export Proposals
# -------------------------
@login_required
def export_proposals_csv(request):
    proposals = Proposal.objects.filter(user=request.user)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="proposals.csv"'

    writer = csv.writer(response)
    writer.writerow(["ID", "Platform", "Status", "Confidence", "Client", "Created At", "Note"])

    for p in proposals:
        writer.writerow([p.id, p.platform, p.status, p.confidence, p.client.name if p.client else "-", p.created_at, p.note])

    return response


@login_required
def export_proposals_excel(request):
    proposals = Proposal.objects.filter(user=request.user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proposals"

    headers = ["ID", "Platform", "Status", "Confidence", "Client", "Created At", "Note"]
    ws.append(headers)

    for p in proposals:
        ws.append([p.id, p.platform, p.status, p.confidence, p.client.name if p.client else "-", str(p.created_at), p.note])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="proposals.xlsx"'
    wb.save(response)
    return response


# -------------------------
# Export Clients
# -------------------------
@login_required
def export_clients_csv(request):
    clients = Client.objects.filter(user=request.user)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="clients.csv"'

    writer = csv.writer(response)
    writer.writerow(["ID", "Name", "Email", "Notes"])

    for c in clients:
        writer.writerow([c.id, c.name, c.email, c.notes])

    return response


@login_required
def export_clients_excel(request):
    clients = Client.objects.filter(user=request.user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"

    headers = ["ID", "Name", "Email", "Notes"]
    ws.append(headers)

    for c in clients:
        ws.append([c.id, c.name, c.email, c.notes])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="clients.xlsx"'
    wb.save(response)
    return response


# -------------------------
# Export Payments
# -------------------------
@login_required
def export_payments_csv(request):
    payments = Payment.objects.filter(proposal__user=request.user)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="payments.csv"'

    writer = csv.writer(response)
    writer.writerow(["ID", "Proposal ID", "Amount", "Status", "Due Date", "Received Date"])

    for p in payments:
        writer.writerow([p.id, p.proposal.id, p.amount, p.status, p.due_date, p.received_date])

    return response


@login_required
def export_payments_excel(request):
    payments = Payment.objects.filter(proposal__user=request.user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = ["ID", "Proposal ID", "Amount", "Status", "Due Date", "Received Date"]
    ws.append(headers)

    for p in payments:
        ws.append([p.id, p.proposal.id, p.amount, p.status, str(p.due_date), str(p.received_date)])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="payments.xlsx"'
    wb.save(response)
    return response


@login_required
def export_all_data_zip(request):
    # Create an in-memory zip
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zf:

        # ----------------- Proposals.csv -----------------
        proposals_io = io.StringIO()
        writer = csv.writer(proposals_io)
        writer.writerow(["ID", "Platform", "Status", "Confidence", "Client", "Created At", "Note"])
        for p in Proposal.objects.filter(user=request.user):
            writer.writerow([
                p.id,
                p.platform,
                p.status,
                p.confidence,
                p.client.name if p.client else "-",
                p.created_at,
                p.note
            ])
        zf.writestr("Proposals.csv", proposals_io.getvalue())
        proposals_io.close()

        # ----------------- Clients.csv -----------------
        clients_io = io.StringIO()
        writer = csv.writer(clients_io)
        writer.writerow(["ID", "Name", "Email", "Notes"])
        for c in Client.objects.filter(user=request.user):
            writer.writerow([c.id, c.name, c.email, c.notes])
        zf.writestr("Clients.csv", clients_io.getvalue())
        clients_io.close()

        # ----------------- Payments.csv -----------------
        payments_io = io.StringIO()
        writer = csv.writer(payments_io)
        writer.writerow(["ID", "Proposal ID", "Amount", "Status", "Due Date", "Received Date"])
        for pay in Payment.objects.filter(proposal__user=request.user):
            writer.writerow([pay.id, pay.proposal.id, pay.amount, pay.status, pay.due_date, pay.received_date])
        zf.writestr("Payments.csv", payments_io.getvalue())
        payments_io.close()

    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/zip")
    response["Content-Disposition"] = 'attachment; filename="all_data_backup.zip"'
    return response


def freelancer_list(request):
    """Public list of freelancers with optional filters."""
    query = request.GET.get("q", "")
    skill_filter = request.GET.get("skill", "")
    profession_filter = request.GET.get("profession", "")

    freelancers = FreelancerDirectoryProfile.objects.filter(is_visible=True)

    if query:
        freelancers = freelancers.filter(display_name__icontains=query)
    if skill_filter:
        freelancers = freelancers.filter(skills__icontains=skill_filter)
    if profession_filter:
        freelancers = freelancers.filter(profession=profession_filter)

    context = {
        "freelancers": freelancers,
        "query": query,
        "skill_filter": skill_filter,
        "profession_filter": profession_filter,
        "profession_choices": FreelancerDirectoryProfile.PROFESSION_CHOICES,
    }
    return render(request, "proposals/freelancer_list.html", context)


def freelancer_detail(request, pk):
    """Public detail page for a freelancer."""
    freelancer = get_object_or_404(FreelancerDirectoryProfile, pk=pk, is_visible=True)
    
    # Prepare skills as a list
    skills_list = []
    if freelancer.skills:
        skills_list = [skill.strip() for skill in freelancer.skills.split(",") if skill.strip()]

    return render(
        request,
        "proposals/freelancer_detail.html",
        {"freelancer": freelancer, "skills_list": skills_list}
    )


@login_required
def edit_freelancer_profile(request):
    """Allow a user to create or update their freelancer directory profile."""
    profile, created = FreelancerDirectoryProfile.objects.get_or_create(user=request.user)

    if request.method == "POST":
        form = FreelancerDirectoryProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            freelancer = form.save(commit=False)
            freelancer.user = request.user
            freelancer.save()
            return redirect("freelancer_detail", pk=freelancer.id)
    else:
        form = FreelancerDirectoryProfileForm(instance=profile)

    return render(request, "proposals/edit_freelancer_profile.html", {"form": form})


# views.py
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib import colors

@login_required
def generate_invoice_pdf(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    payments = proposal.payments.all().order_by("due_date")

    # PDF response
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="invoice_{proposal_id}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter,
                            rightMargin=40, leftMargin=40, topMargin=50, bottomMargin=50)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], alignment=TA_CENTER)
    elements.append(Paragraph("Invoice", title_style))
    elements.append(Spacer(1, 20))

    # Proposal & Client Info
    client_info = f"""
    <b>Client:</b> {proposal.client.name if proposal.client else "N/A"}<br/>
    <b>Email:</b> {proposal.client.email if proposal.client else "N/A"}<br/>
    <b>Project:</b> {proposal.job_description[:80]}...
    """
    elements.append(Paragraph(client_info, styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Table of Payments
    data = [["Due Date", "Received Date", "Status", "Amount (USD)"]]
    for p in payments:
        data.append([
            p.due_date.strftime("%Y-%m-%d") if p.due_date else "-",
            p.received_date.strftime("%Y-%m-%d") if p.received_date else "-",
            p.status,
            f"${p.amount}"
        ])

    table = Table(data, colWidths=[100, 100, 100, 100])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
        ("ALIGN", (-1,0), (-1,-1), "RIGHT"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 20))

    # Total
    total_amount = sum(p.amount for p in payments)
    elements.append(Paragraph(f"<b>Total Amount:</b> ${total_amount}", styles["Normal"]))

    # Build PDF
    doc.build(elements)
    return response


def terms(request):
    return render(request, "proposals/terms.html")

def privacy(request):
    return render(request, "proposals/privacy.html")

def contact(request):
    return render(request, "proposals/contact.html")

def about(request):
    return render(request, "proposals/about.html")


@login_required
def credits_overview(request):
    profile = Profile.objects.get(user=request.user)
    txns = CreditTransaction.objects.filter(user=request.user).order_by("-created_at")[:50]
    return render(request, "billing/credits_overview.html", {
        "balance": profile.ai_credits,
        "transactions": txns,
    })

@user_passes_test(lambda u: u.is_staff)
def credits_admin_adjust(request):
    if request.method == "POST":
        form = AdminAdjustCreditsForm(request.POST)
        if form.is_valid():
            user = form.cleaned_data["user"]
            delta = form.cleaned_data["delta"]
            reason = form.cleaned_data["reason"]
            method = form.cleaned_data["method"]
            note = form.cleaned_data["note"]

            try:
                new_balance = adjust_credits(user, delta, reason, method=method, note=note, created_by=request.user)
                messages.success(request, f"Adjusted {user.username}: {delta} credits → new balance {new_balance}.")
            except ValueError as e:
                messages.error(request, str(e))
            return redirect("credits_admin_adjust")
    else:
        form = AdminAdjustCreditsForm()
    return render(request, "billing/credits_admin_adjust.html", {"form": form})


@login_required
def credits_request(request):
    if request.method == "POST":
        form = CreditRequestForm(request.POST)
        if form.is_valid():
            amount = form.cleaned_data["amount"]
            message_text = form.cleaned_data["message"]

            # ✅ Save to DB
            CreditRequest.objects.create(
                user=request.user,
                amount_requested=amount,
                message=message_text,
                status="pending"   # default anyway, but explicit
            )

            messages.success(request, f"Request received for {amount} credits. Our team will review it shortly.")
            return redirect("credits_overview")
    else:
        form = CreditRequestForm()
    return render(request, "billing/credits_request.html", {"form": form})



from .models import CreditRequest

@login_required
@user_passes_test(lambda u: u.is_staff)
def credit_requests_list(request):
    requests = CreditRequest.objects.filter(status="pending").select_related("user")
    return render(request, "billing/credit_requests_list.html", {"requests": requests})

@login_required
@user_passes_test(lambda u: u.is_staff)
def credit_request_action(request, request_id, action):
    credit_request = get_object_or_404(CreditRequest, id=request_id)

    if action == "approve" and credit_request.status == "pending":
        try:
            adjust_credits(
                credit_request.user,
                credit_request.amount_requested,
                reason="manual_topup",
                method="manual",
                note=f"Approved credit request ID {credit_request.id}",
                created_by=request.user,
            )
            credit_request.status = "approved"
            messages.success(request, f"Approved {credit_request.user} request for {credit_request.amount_requested} credits.")
        except Exception as e:
            messages.error(request, f"Error updating credits: {e}")
            return redirect("credit_requests_list")

    elif action == "reject" and credit_request.status == "pending":
        credit_request.status = "rejected"
        messages.info(request, f"Rejected {credit_request.user}'s request.")

    credit_request.save()
    return redirect("credit_requests_list")


@login_required
def add_conversation(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)

    if request.method == "POST":
        sender = request.POST.get("sender")
        message = request.POST.get("message")

        if message.strip():
            Conversation.objects.create(
                proposal=proposal,
                sender=sender,
                message=message
            )
            messages.success(request, "Conversation added successfully.")
            return redirect("track_conversations", proposal_id=proposal.id)

    return render(request, "proposals/add_conversation.html", {"proposal": proposal})


@login_required
def track_conversations(request, proposal_id):
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    conversations = Conversation.objects.filter(proposal=proposal)

    return render(request, "proposals/track_conversations.html", {
        "proposal": proposal,
        "conversations": conversations
    })


@login_required
def download_conversations_pdf(request, proposal_id):
    """Generate and return a PDF of all conversations for a proposal."""
    proposal = get_object_or_404(Proposal, id=proposal_id, user=request.user)
    conversations = Conversation.objects.filter(proposal=proposal).order_by("date")

    # Create PDF in memory
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Header
    p.setFont("Helvetica-Bold", 16)
    p.drawString(1 * inch, height - 1 * inch, f"Conversations for Proposal #{proposal.id}")
    p.setFont("Helvetica", 12)
    p.drawString(1 * inch, height - 1.3 * inch, f"Job: {proposal.job_description[:80]}")
    p.line(1 * inch, height - 1.4 * inch, width - 1 * inch, height - 1.4 * inch)

    y = height - 1.7 * inch
    p.setFont("Helvetica", 10)

    for convo in conversations:
        text = f"[{convo.date.strftime('%Y-%m-%d %H:%M')}] {convo.sender}: {convo.message}"
        
        # Wrap text if it's too long
        max_width = width - 2 * inch
        lines = []
        while len(text) > 0:
            if p.stringWidth(text, "Helvetica", 10) <= max_width:
                lines.append(text)
                text = ""
            else:
                split_at = text.rfind(" ", 0, 90)  # break at word boundary
                if split_at == -1:
                    split_at = 90
                lines.append(text[:split_at])
                text = text[split_at:].lstrip()

        for line in lines:
            if y < 1 * inch:  # New page if too low
                p.showPage()
                p.setFont("Helvetica", 10)
                y = height - 1 * inch
            p.drawString(1 * inch, y, line)
            y -= 14  # line spacing

        y -= 8  # small gap between messages

    p.showPage()
    p.save()

    # Prepare response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="conversations_proposal_{proposal.id}.pdf"'
    return response

from .forms import ProgressReportRequestForm
from .models import ProgressReport

@login_required
def request_progress_report(request):
    if request.method == "POST":
        form = ProgressReportRequestForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.user = request.user
            report.status = "pending"
            report.save()
            messages.success(request, "✅ Your request has been submitted! Please complete the payment to proceed.")
            return redirect("my_requests")
    else:
        form = ProgressReportRequestForm()

    return render(request, "proposals/request_progress_report.html", {"form": form})

@staff_member_required
def progress_report_requests_list(request):
    requests = ProgressReport.objects.filter(status="pending").select_related("user")
    return render(request, "proposals/progress_report_requests_list.html", {"requests": requests})

@staff_member_required
def admin_generate_progress_report(request, report_id):
    report = get_object_or_404(ProgressReport, id=report_id, status="pending")

    # For now just mark as processing (later you can hook in your LLM logic)
    report.status = "processing"
    report.save()

    messages.success(request, f"Started processing report for {report.user.username}.")
    return redirect("progress_report_requests_list")

@staff_member_required
def admin_reject_progress_report(request, report_id):
    report = get_object_or_404(ProgressReport, id=report_id, status="pending")
    report.status = "rejected"
    report.save()

    messages.info(request, f"Rejected progress report request from {report.user.username}.")
    return redirect("progress_report_requests_list")

@login_required
def my_requests(request):
    credit_requests = CreditRequest.objects.filter(user=request.user).order_by("-created_at")
    progress_reports = ProgressReport.objects.filter(user=request.user).order_by("-created_at")

    return render(request, "proposals/my_requests.html", {
        "credit_requests": credit_requests,
        "progress_reports": progress_reports,
    })

@user_passes_test(lambda u: u.is_staff)
def admin_workspace(request):
    credit_requests = CreditRequest.objects.filter(status="pending").select_related("user")
    progress_requests = ProgressReport.objects.filter(status="pending").select_related("user")

    return render(request, "proposals/admin_workspace.html", {
        "credit_requests": credit_requests,
        "progress_requests": progress_requests,
    })


from django.utils.timezone import make_aware

@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_generate_progress_report(request, report_id):
    report = get_object_or_404(ProgressReport, id=report_id)

    # 1️⃣ Build query filters
    start_date = make_aware(datetime.datetime(report.year, report.month, 1))
    if report.month == 12:
        end_date = make_aware(datetime.datetime(report.year + 1, 1, 1))
    else:
        end_date = make_aware(datetime.datetime(report.year, report.month + 1, 1))

    proposals = Proposal.objects.filter(
        user=report.user,
        created_at__range=(start_date, end_date)
    )

    if report.platform != "all":
        proposals = proposals.filter(platform=report.platform)
    if report.status_filter != "all":
        proposals = proposals.filter(status=report.status_filter)

    conversations = Conversation.objects.filter(proposal__in=proposals)

    # 2️⃣ Build context for LLM
    context = []
    for p in proposals:
        context.append({
            "job_title": p.job_title,
            "platform": p.platform,
            "status": p.status,
            "proposal_text": p.proposal_text,
            "conversations": list(conversations.filter(proposal=p).values("sender", "message", "date")),
        })

    # 3️⃣ Send to LLM for analysis
    from openai import OpenAI
    client = OpenAI()
    prompt = f"""
    Analyze the following user proposals and conversations.
    For each project, summarize:
    - What went well
    - What was the deal breaker (if lost)
    - Key USPs or strengths
    - Skills mismatch or improvement points
    - Suggestions for approaching similar clients

    User data:
    {context}
    """

    llm_response = client.chat.completions.create(
        model="gpt-4o",  # or any model you prefer
        messages=[
            {"role": "system", "content": "You are a career coach analyzing freelancing conversations and proposals."},
            {"role": "user", "content": prompt}
        ]
    )

    report_text = llm_response.choices[0].message.content

    # 4️⃣ Save as text or PDF
    from io import BytesIO
    from reportlab.platypus import SimpleDocTemplate, Paragraph
    from reportlab.lib.styles import getSampleStyleSheet

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer)
    styles = getSampleStyleSheet()
    story = [Paragraph(report_text, styles["BodyText"])]
    doc.build(story)

    report.generated_report.save(f"progress_report_{report.user.username}_{report.month}_{report.year}.pdf", buffer)
    report.status = "Generated"
    report.save()

    messages.success(request, f"Progress report generated for {report.user.username}.")
    return redirect("progress_report_requests_list")
